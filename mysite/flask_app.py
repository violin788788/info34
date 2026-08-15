import os
from flask import Flask, render_template, request, jsonify
from openpyxl import load_workbook
app = Flask(__name__)
XLSX_PATH = os.path.join(app.root_path, 'books.xlsx')
@app.route('/')
def home():
    static_folder = os.path.join(app.root_path, 'static')
    mp3_files = []
    if os.path.exists(static_folder):
        mp3_files = [f for f in os.listdir(static_folder) if f.endswith('.mp3')]
    return render_template('info34.html', mp3_files=mp3_files)
@app.route('/books2')
def books():
    wb = load_workbook(XLSX_PATH, data_only=False)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append([cell.value if cell.value is not None else '' for cell in row])
    wb.close()
    return render_template('books2.html', rows=rows)
@app.route('/update_book', methods=['POST'])
def update_book():
    data = request.get_json(silent=True)
    if data is None:
        data = request.form
    row = data.get('row')
    column = data.get('column')
    value = data.get('value')
    if row is None or column is None:
        return jsonify({'success': False, 'error': 'Missing row or column'}), 400
    try:
        row = int(row)
        column = int(column)
        wb = load_workbook(XLSX_PATH, data_only=False)
        ws = wb.active
        ws.cell(row=row + 1, column=column + 1).value = value
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(XLSX_PATH)
        wb.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
if __name__ == '__main__':
    app.run(debug=True)




"""
import os
import sqlite3
from flask import Flask, render_template, request, jsonify
app = Flask(__name__)
DB_PATH = os.path.join(app.root_path, 'books.db')
@app.route('/')
def home():
    static_folder = os.path.join(app.root_path, 'static')
    mp3_files = []
    if os.path.exists(static_folder):
        mp3_files = [f for f in os.listdir(static_folder) if f.endswith('.mp3')]
    return render_template('info34.html', mp3_files=mp3_files)
@app.route('/books')
def books():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.execute('SELECT * FROM books')
    columns = [description[0] for description in cursor.description]
    books = cursor.fetchall()
    conn.close()
    return render_template('books.html', books=books, columns=columns)
@app.route('/update_book', methods=['POST'])
def update_book():
    data = request.get_json()
    book_id = data.get('id')
    field = data.get('field')
    value = data.get('value')
    if not book_id or not field:
        return jsonify({'success': False, 'error': 'Missing ID or field'}), 400
    conn = sqlite3.connect(DB_PATH)
    columns = [row[1] for row in conn.execute('PRAGMA table_info(books)').fetchall()]
    if field not in columns:
        conn.close()
        return jsonify({'success': False, 'error': 'Invalid field'}), 400
    if field == 'id':
        conn.close()
        return jsonify({'success': False, 'error': 'ID cannot be edited'}), 400
    safe_field = '"' + field.replace('"', '""') + '"'
    try:
        conn.execute(f'UPDATE books SET {safe_field} = ? WHERE id = ?', (value, book_id))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
if __name__ == '__main__':
    app.run(debug=True)



"""